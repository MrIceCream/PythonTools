
import xml.etree.ElementTree as ET
import openpyxl

file_path = '/Volumes/T7/cook/cook-client-sunshine/Assets/ChefMenu/Resources/Configs/Configs/DefsConfig.xml'
city_config_path = '/Volumes/T7/cook/cook-client-sunshine/Assets/ChefMenu/Resources/Configs/Balance'

def get_all_cities():
    all_city = []
    tree = ET.parse(file_path)
    root = tree.getroot()
    cities = root.find('auto').find('cities')
    for city in cities:
        city_name = city.attrib['name']
        all_city.append(city_name)
    return all_city

def get_level_map():
    level_map = {}
    cities = get_all_cities()
    for city in cities:
        tree = ET.parse(f'{city_config_path}/CityConfig_{city}.xml')
        root = tree.getroot().find('city')
        total_level = 0
        region_index = 0
        for region in root.findall('region'):
            levels = region.find('levels').findall('level')
            key = f'{city}_{region_index}'
            level_map[key] = total_level

            total_level += len(levels)
            region_index += 1
    return level_map

def deal_xlsx():
    level_map = get_level_map()
    # 打开Excel文件
    workbook = openpyxl.load_workbook('billboards.xlsx')
    # 选择要操作的表格（如果有多个表格，根据实际情况选择）
    sheet = workbook.active
    # 遍历每一行，并进行相加操作
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    # for row in sheet.iter_rows(min_row=2, values_only=True):
        key = f'{row[0]}_{row[1]}'
        sheet.cell(row=index, column=6).value = row[3] + level_map[key]
        sheet.cell(row=index, column=7).value = row[4] + level_map[key]

    # 保存修改后的Excel文件
    workbook.save('billboards_deal.xlsx')

deal_xlsx()