#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import os

replace_list = [
    {"target":"apollo","source":"gramgames"},
    {"target":"Apollo","source":"Gramgames"},
    {"target":"Apollo","source":"Gram Games"},
    {"target":"apollo","source":"gram"},
    {"target":"Servoapollo","source":"Servogram"},
    {"target":"servoapollo","source":"servogram"},
    {"target":"SERVOAPOLLO","source":"SERVOGRAM"},
    {"target":"Instoapollo","source":"Instogram"},
    {"target":"instoapollo","source":"instogram "},
    {"target":"INSTOAPOLLO","source":"INSTOGRAM"},
    {"target":"fusion","source":"mergemagic"},
    {"target":"Fusion","source":"MergeMagic"},
    {"target":"Fusion","source":"Merge Magic!"},
    {"target":"Fusion","source":"Merge Magic"},
    {"target":"fusionbase","source":"mergedragons"},
    {"target":"FusionBase","source":"MergeDragons"},
    {"target":"tortoise","source":"dragon"},
    {"target":"Tortoise","source":"Dragon"},
    {"target":"TORTOISE","source":"DRAGON"},
    {"target":"1456716201","source":"1462419002"}
]


def get_fullpath_by_extension(rootdir, extensions=None):
    ''' 获取路径下的所有文件 根据extensions '''
    __re = []
    for parent, dirnames, filenames in os.walk(rootdir):
        for filename in filenames:
            fullpath = os.path.join(parent, filename)
            path_extension = os.path.splitext(fullpath)[1]
            if (extensions and path_extension in extensions) or not extensions:
                result = fullpath.replace("\\", "/")
                __re.append(result)
    return __re


def replaceXlsxContentString(file):
    workbook = load_workbook(file)
    sheetnames = workbook.sheetnames

    print('='*50)
    print('source file: '+file)
    for sheetname in sheetnames:
        print('source sheet: '+sheetname)
        oldsheetname = sheetname
        sheet = workbook[sheetname]

        # 校验子表名
        change = False
        for replace_item in replace_list:
            if sheetname.find(replace_item['source']) != -1:
                sheetname = sheetname.replace(
                    replace_item['source'], replace_item['target'])
                change = True

        if change:
            print('old sheet: {old}, new sheet: {new}'.format(
                old=oldsheetname, new=sheetname))
            sheet.title = sheetname

        rows = sheet.rows
        for row in rows:
            for col in row:
                change = False
                old_value = col.value
                if old_value == None or not isinstance(old_value, str):
                    continue
                
                # col.number_format = '@'
                
                for replace_item in replace_list:
                    if col.value.find(replace_item['source']) != -1:
                        col.value = col.value.replace(
                            replace_item['source'], replace_item['target'])
                        change = True

                if change:
                    print('old : {old}, new : {new}'.format(
                        old=old_value, new=col.value))
    

    target_file = file.replace('input','output')
    # print('target file: '+target_file)
    parent_dir = os.path.dirname(target_file)
    if not os.path.exists(parent_dir):
        os.makedirs(parent_dir)
    workbook.save(target_file);



def main():
    excel_path = os.path.join(os.getcwd(), 'input')
    file_list = get_fullpath_by_extension(excel_path, ['.xlsx'])
    for file in file_list:
        replaceXlsxContentString(file)


if __name__ == "__main__":
    main()
