import openpyxl
from openpyxl.styles import Alignment
import random

def create_layer(layer):
    layer_list = str(layer).split('.')
    if layer != 1:
        layer_list.append("1")
    return '.'.join(layer_list)

def increase_layer(layer):
    layer_list = str(layer).split('.')
    layer_list[-1] = str(int(layer_list[-1]) + 1)
    return '.'.join(layer_list)

def main():
    workbook = openpyxl.load_workbook('模板.xlsx')
    sheet = workbook.active

    insert_col = 4
    insert_amount = 4

    sheet.insert_cols(insert_col, amount=insert_amount)
    sheet.cell(row=1, column=insert_col, value="新层级")
    sheet.cell(row=1, column=insert_col+1, value="模拟库存")
    sheet.cell(row=1, column=insert_col+2, value="齐套使用量")
    sheet.cell(row=1, column=insert_col+3, value="齐套后剩余库存")

    pre_value = 0
    index_map = {0:1}
    bom = {'min_unit':0, 'inventory':0, 'components':{}}

    for row in sheet.iter_rows(min_row=2):
        row_number = row[0].row
        layer = row[2].value
        value = int( layer.replace('-', '').replace(' ', '') )
        if value >= pre_value:
            if value not in index_map:
                index_map[value] = create_layer(index_map[pre_value])
            else:
                index_map[value] = increase_layer(index_map[value])
            new_value = index_map[value]
        else:
            del index_map[pre_value]
            index_map[value] = increase_layer(index_map[value])
            new_value = index_map[value]
        pre_value = value
        # 添加新层级
        sheet.cell(row=row_number, column=insert_col, value=new_value)

        min_unit = row[12+insert_amount].value 

        # 添加随机库存
        mock_inventory = int( random.randrange(4, 10) * min_unit)
        sheet.cell(row=row_number, column=insert_col+1, value=mock_inventory)

        # Bom树
        tree = new_value.split('.')
        parent = bom
        key_list = []
        for item in tree:
            key_list.append(item)
            key = '.'.join(key_list)

            if key not in parent['components']:
                parent['components'][key] = {'min_unit':min_unit, 'inventory':mock_inventory, 'components':{}}
                if item == '1':
                    parent['inventory'] = 0
            parent = parent['components'][key]

        if len(tree) > 1 and item == '1':
            sheet.cell(row=row_number-1, column=insert_col+1, value=0)

    # 齐套计算
    suit = calculate_total_quantity(bom, 'root')
    print('齐套数:{0}'.format(suit))

    # 齐套数据写入表格
    for row in sheet.iter_rows(min_row=2):
        row_number = row[0].row
        layer = row[3].value
        # Bom树
        tree = layer.split('.')
        this_bom = bom
        key_list = []
        for item in tree:
            key_list.append(item)
            key = '.'.join(key_list)
            this_bom = this_bom['components'][key]

        # 添加随机库存齐套
        sheet.cell(row=row_number, column=insert_col+2, value= this_bom['suit_use'])
        # 添加齐套后剩余库存
        sheet.cell(row=row_number, column=insert_col+3, value= this_bom['remain_inv'])

    workbook.save('结果.xlsx')  # 替换为你想要保存的文件名


# 递归累加计算到最顶层的数量
def calculate_total_quantity(bom, parent_key):
    inventory = bom['inventory']
    components = bom['components']
    
    if len(components) == 0:
        return inventory

    total = inventory
    min_suit = -1
    for key in components:
        component = components[key]
        component['inventory'] = calculate_total_quantity(component, key)
        avaliable_suit = (int) (component['inventory'] / component['min_unit'])

        if avaliable_suit == 0:
            print('{0} 齐套为0'.format(key))
        # 初始化
        min_suit = avaliable_suit if min_suit == -1 else min_suit
        # 取最小齐套值
        min_suit = min_suit if min_suit < avaliable_suit else avaliable_suit

    # 插入齐套数据
    for key in components:
        component = components[key]
        component['suit_use'] = min_suit * component['min_unit']
        component['remain_inv'] = component['inventory'] - component['suit_use']

    total = total + min_suit
    return total

main()