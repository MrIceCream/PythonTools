from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from PIL import Image as PILImage, ImageFile
import os
import pandas as pd

csv_folder = "csv"
xlsx_folder = "xlsx"
root_path = "/Volumes/T7/cook/cook-client-sunshine/"


def deal_city_kitchen_image(xlsx_file):
    wb = load_workbook(xlsx_file)
    ws = wb.active
    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[4].value == None:
            continue
        img_path = os.path.join(root_path, row[4].value)
        img_name = os.path.basename(img_path)
        # new_image_path = "img/{0}".format(img_name)
        new_image_path = "img/{0}/{1}".format(row[0].value, img_name)

        # 图片压缩
        if not os.path.exists(new_image_path):
            try:
                ImageFile.LOAD_TRUNCATED_IMAGES = True
                original_image = PILImage.open(img_path)
            except Exception as e:
                print("打开图片错误, {0}, {1}".format(img_path, e))
                continue
            width, height = original_image.size
            if width > 1 and height > 1:
                new_width = int(width * 0.5)
                new_height = int(height * 0.5)
                original_image = original_image.resize((new_width, new_height))
            os.makedirs(os.path.dirname(new_image_path), exist_ok=True)
            original_image.save(new_image_path, quality=85, optimize=True)

        img = Image(new_image_path)
        # img = Image(img_path)
        ws.row_dimensions[row_number].height = img.height
        ws.add_image(img, anchor="D" + str(row_number))

        # row[4].value = os.path.basename(row[4].value)
        row[4].value = row[4].value.replace("Assets/ChefMenu/Images/Kitchens/", "")
        if row[5].value != None:
            row[5].value = row[5].value.replace("Assets/ChefMenu/Prefabs/Kitchens/", "")

        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自适应调整单元格的宽度
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)

        if column[0].column == 4:
            ws.column_dimensions[column_letter].width = 100
        else:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save("output/" + os.path.basename(xlsx_file))


def deal_city_kitchen_table_image(xlsx_file):
    wb = load_workbook(xlsx_file)
    ws = wb.active
    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        img_path = os.path.join(root_path, row[2].value)
        img_name = os.path.basename(img_path)
        # new_image_path = "img/{0}".format(img_name)
        new_image_path = "img/{0}/{1}".format(row[0].value, img_name)

        # 图片压缩
        if not os.path.exists(new_image_path):
            try:
                ImageFile.LOAD_TRUNCATED_IMAGES = True
                original_image = PILImage.open(img_path)
            except Exception as e:
                print("打开图片错误, {0}, {1}".format(img_path, e))
                continue
            width, height = original_image.size
            if width > 1 and height > 1:
                new_width = int(width * 0.5)
                new_height = int(height * 0.5)
                original_image = original_image.resize((new_width, new_height))
            os.makedirs(os.path.dirname(new_image_path), exist_ok=True)
            original_image.save(new_image_path, quality=85, optimize=True)

        img = Image(new_image_path)
        # img = Image(img_path)
        ws.row_dimensions[row_number].height = img.height
        ws.add_image(img, anchor="B" + str(row_number))

        row[2].value = row[2].value.replace("Assets/ChefMenu/Images/Kitchens/", "")

        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自适应调整单元格的宽度
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)

        if column[0].column == 2:
            ws.column_dimensions[column_letter].width = 100
        else:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save("output/" + os.path.basename(xlsx_file))


def main():
    for filename in os.listdir(xlsx_folder):
        if filename.endswith(".xlsx"):
            xlsx_file = os.path.join(xlsx_folder, filename)
            # deal_city_kitchen_image(xlsx_file)
            deal_city_kitchen_table_image(xlsx_file)


def csv_2_xlsx():
    if not os.path.exists(xlsx_folder):
        os.makedirs(xlsx_folder)

    for filename in os.listdir(csv_folder):
        if filename.endswith(".csv"):
            csv_file = os.path.join(csv_folder, filename)
            data = pd.read_csv(csv_file)
            xlsx_file = os.path.join(
                xlsx_folder, os.path.splitext(filename)[0] + ".xlsx"
            )
            writer = pd.ExcelWriter(xlsx_file, engine="xlsxwriter")
            data.to_excel(writer, index=False, sheet_name="Sheet1")
            writer.close()


if __name__ == "__main__":
    # csv_2_xlsx()
    main()
