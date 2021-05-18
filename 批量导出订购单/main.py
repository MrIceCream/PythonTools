# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def dealSource():
    wb = load_workbook('input/数据源.xlsx')
    ws = wb.active

    resultDict = {}
    for row in ws.iter_rows(min_row=2):
        company = row[2].value
        if not resultDict.__contains__(company):
            resultDict[company] = {}
        companyDetail = resultDict[company]

        orderId = row[0].value
        if not companyDetail.__contains__(orderId):
            companyDetail[orderId] = {
                '订购编号': row[0].value,
                '合同编号': row[1].value,
                '乙方': row[2].value,
                '联系人': row[3].value,
                '联系方式': row[4].value,
                '付款方式': row[5].value,
                '日期': row[13].value
            }
        orderDetail = companyDetail[orderId]

        if not orderDetail.__contains__('SKUS'):
            orderDetail['SKUS'] = []

        skuDetail = {
            'SKU': row[6].value,
            '货物名称': row[7].value,
            '单位': row[8].value,
            '数量': row[10].value,
            '含税单价': row[11].value,
            '总金额': row[12].value,
            '币种': 'CNY',
            '交货日期': row[9].value,
        }
        orderDetail['SKUS'].append(skuDetail)

    return resultDict

def generateResult(sourceData):
    for company in sourceData:
        tempalteWb = load_workbook('input/template.xlsx')
        ws = tempalteWb['Table 1']
        companyDetail = sourceData[company];
        for orderId in companyDetail:
            orderDetail = companyDetail[orderId];
            orderWs = tempalteWb.copy_worksheet(ws)
            orderWs.title = orderId
            orderWs['B5'] = company
            orderWs['B6'] = orderDetail['联系人']
            orderWs['B7'] = orderDetail['合同编号']
            orderWs['B8'] = orderDetail['订购编号']
            orderWs['F5'] = orderDetail['日期']
            orderWs['F6'] = orderDetail['联系方式']
            orderWs['F7'] = orderDetail['付款方式']

            skuLens = len(orderDetail['SKUS'])
            # deleteRows = 20-skuLens
            # orderWs.delete_rows(10, deleteRows)

            totalCount = 0
            totalPrice = 0
            start_row = 10
            for i in range(0, skuLens):
                skuDetail = orderDetail['SKUS'][i]
                orderWs['A'+str(start_row+i)] = skuDetail['SKU']
                orderWs['B'+str(start_row+i)] = skuDetail['货物名称']
                orderWs['C'+str(start_row+i)] = skuDetail['单位']
                orderWs['D'+str(start_row+i)] = skuDetail['数量']
                orderWs['E'+str(start_row+i)] = skuDetail['含税单价']
                orderWs['F'+str(start_row+i)] = skuDetail['总金额']
                orderWs['G'+str(start_row+i)] = skuDetail['币种']
                orderWs['H'+str(start_row+i)] = skuDetail['交货日期']
                totalCount += skuDetail['数量']
                totalPrice += skuDetail['总金额']

            skuLens = 20
            totalPrice = round(totalPrice,2)
            orderWs['D'+str(10+skuLens)] = totalCount
            orderWs['F'+str(10+skuLens)] = totalPrice
            orderWs['F8'] = digital_to_chinese(totalPrice)
            orderWs['B'+str(24+skuLens)] = orderDetail['日期']
            orderWs['F'+str(22+skuLens)] = orderDetail['乙方']
            orderWs['F'+str(24+skuLens)] = orderDetail['日期']
        
        tempalteWb.remove(ws)
        tempalteWb.save('output/'+company+'.xlsx')

def digital_to_chinese(digital):
    str_digital = str(digital)
    chinese = {'1': '壹', '2': '贰', '3': '叁', '4': '肆', '5': '伍', '6': '陆', '7': '柒', '8': '捌', '9': '玖', '0': '零'}
    chinese2 = ['拾', '佰', '仟', '万', '厘', '分', '角']
    jiao = ''
    bs = str_digital.split('.')
    yuan = bs[0]
    if len(bs) > 1:
        jiao = bs[1]
    r_yuan = [i for i in reversed(yuan)]
    count = 0
    for i in range(len(yuan)):
        if i == 0:
            r_yuan[i] += '元'
            continue
        r_yuan[i] += chinese2[count]
        count += 1
        if count == 4:
            count = 0
            chinese2[3] = '亿'

    s_jiao = [i for i in jiao][:3]  # 去掉小于厘之后的

    j_count = -1
    for i in range(len(s_jiao)):
        s_jiao[i] += chinese2[j_count]
        j_count -= 1
    last = [i for i in reversed(r_yuan)] + s_jiao

    last_str = ''.join(last)
    print(str_digital)
    print(last_str)
    last_str = last_str.replace('0拾', '0').replace('0佰', '0').replace('0仟', '0').replace('0万', '万').replace('00', '0').replace('000', '0').replace('0元', '元')
    for i in range(len(last_str)):
        digital = last_str[i]
        if digital in chinese:
            last_str = last_str.replace(digital, chinese[digital])
    
    if jiao == '':
        last_str += '整'

    print(last_str)
    return last_str

def main():
    resultDict = dealSource()
    generateResult(resultDict)

main()