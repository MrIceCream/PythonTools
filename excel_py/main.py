#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook


INVOICE_TEMPLATE_PATH = './template/invoice_template.xlsx'
PROCUREMENT_PATH = './input/procurement.xlsx'


DELIVERY_TEMPLATE_PATH = './template/delivery_template.xlsx'
RECEVING_PATH = './input/receiving.xlsx'


'''
读取excel （删除第一行表头）
'''
def read_excel(path):
    result = []

    workbook = load_workbook(path)
    sheet = workbook.active
    rows = sheet.rows
    for row in rows:
        line = [col.value for col in row]
        result.append(line)

    result.pop(0)

    return result


'''
导出发票信息
'''
def export_invoice():
    excel_date = read_excel(PROCUREMENT_PATH)

    cg_dict = {}
    for row in excel_date:
        cg_key = row[0]
        if cg_key in cg_dict:
            cg_dict[cg_key].append(row)
        else:
            cg_dict[cg_key] = [row]


    for cg_key in cg_dict:
        cg_data = cg_dict[cg_key]
        sku_dict = deal_procurement_cg_data(cg_data)
        write_invoice_file(cg_key, sku_dict)
        print(cg_key, len(sku_dict))


'''
处理采购单同一cg信息
'''
def deal_procurement_cg_data(cg_data):
    sku_dict = {}
    for row in cg_data:
        sku_key = row[1]
        if sku_key in sku_dict:
            sku_dict[sku_key]["count"] += row[3]
            sku_dict[sku_key]["total"] += row[6]
        else:
            sku_dict[sku_key] = {"name":    row[8],
                                 "count":   row[3],
                                 "unit":    row[2],
                                 "price":   row[5],
                                 "total":   row[6],
                                 "sku":     row[1]
                                 }
        sku_dict["date"] = row[4]
        sku_dict['cg_total_price'] = row[7]

    return sku_dict


'''
导出发票excel
'''
def write_invoice_file(cg_key, sku_dict):
    
    invoice_template_empty_row_index = 8
    invoice_template_empty_row = 25

    workbook = load_workbook(INVOICE_TEMPLATE_PATH)
    worksheet = workbook.active
    worksheet.title = u'发票'

    worksheet.cell(row=3, column=1, value='Contract No.：'+str(cg_key))
    cg_time = sku_dict['date'].replace('-', '/').split(' ')[0]
    worksheet.cell(row=6, column=6, value=cg_time)

    sku_dict_len = len(sku_dict)-2 # 扣除日期，价格

    worksheet.delete_rows(invoice_template_empty_row_index,
                          invoice_template_empty_row - sku_dict_len - 2) # 空余两行

    count_sum='=SUM(B8:B{end_row})'.format(end_row=invoice_template_empty_row_index + sku_dict_len)
    price_sum='=SUM(E8:E{end_row})'.format(end_row=invoice_template_empty_row_index + sku_dict_len)

    worksheet.cell(row=invoice_template_empty_row_index+sku_dict_len+2, column=2, value=count_sum)
    worksheet.cell(row=invoice_template_empty_row_index+sku_dict_len+2, column=5, value=price_sum)

    row = invoice_template_empty_row_index
    total_price = 0
    for sku_key in sku_dict:

        if sku_key == "date" or sku_key == 'cg_total_price':
            continue

        sku_data = sku_dict[sku_key]
        worksheet.cell(row=row, column=1).value = sku_data["name"]
        worksheet.cell(row=row, column=2).value = sku_data["count"]
        worksheet.cell(row=row, column=3).value = sku_data["unit"]
        worksheet.cell(row=row, column=4).value = sku_data["price"]
        worksheet.cell(row=row, column=5).value = sku_data["total"]
        worksheet.cell(row=row, column=6).value = sku_data["sku"]

        if sku_data["count"] * sku_data["price"] - sku_data["total"] > 0.01:
            print('商品价格不正确')

        row += 1
        total_price += sku_data["total"]

    if total_price - sku_dict['cg_total_price'] > 0.01:
        print('总价格不正确')

    workbook.save('./output/invoice/'+cg_key+u'发票.xlsx')


'''
导出送货单信息
'''
def export_delivery():
    excel_date = read_excel(RECEVING_PATH)

    cg_dict = {}
    for row in excel_date:
        cg_key = row[5]
        if cg_key in cg_dict:
            cg_dict[cg_key].append(row)
        else:
            cg_dict[cg_key] = [row]


    for cg_key in cg_dict:
        cg_data = cg_dict[cg_key]
        cg_month_dict = deal_receiving_cg_data(cg_data)

        for month in cg_month_dict:
            write_delivery_file(cg_key, month, cg_month_dict[month])

        # break

'''
处理收货单同一cg信息
'''
def deal_receiving_cg_data(cg_data):
    month_dict = {}
    for row in cg_data:
        receiving_month = row[6].split('-')[1]
        receiving_day = int(row[6].split(' ')[0].split('-')[2])

        receiving_single = {    "name":         row[2],
                                "unit":         row[3],
                                "count":        row[4],
                                "real_count":   row[4],
                                "sku":          row[1],
                                "order":        row[0]}

        if receiving_month in month_dict:
            if receiving_day < month_dict[receiving_month]['day']:
                month_dict[receiving_month]['day'] = receiving_day
            month_dict[receiving_month]['receiving_list'].append(receiving_single)
        else:
            month_dict[receiving_month] = {"day": receiving_day, "receiving_list": [receiving_single]}


    return month_dict


'''
处理收货单同一cg信息（合并相同收货单，相同sku，相同cg）
'''
def deal_receiving_cg_data_merge(cg_data):
    month_dict = {}
    for row in cg_data:
        receiving_month = row[6].split('-')[1]
        receiving_day = int(row[6].split(' ')[0].split('-')[2])

        receiving_single = {    "name":         row[2],
                                "unit":         row[3],
                                "count":        row[4],
                                "real_count":   row[4],
                                "sku":          row[2],
                                "order":        row[1]}

        if receiving_month in month_dict:
            if receiving_day < month_dict[receiving_month]['day']:
                month_dict[receiving_month]['day'] = receiving_day
            month_dict[receiving_month]['receiving_list'].append(receiving_single)
        else:
            month_dict[receiving_month] = {"day": receiving_day, "receiving_list": [receiving_single]}


    return month_dict



'''
导出送货单excel
'''
def write_delivery_file(cg_key, month, cg_month_dict):
    delivery_template_empty_row_index = 5
    delivery_template_empty_row = 148

    workbook = load_workbook(DELIVERY_TEMPLATE_PATH)
    worksheet = workbook.active
    worksheet.title = u'送货单'

    worksheet.cell(row=2, column=1, value=u'合同编号：' + str(cg_key))
    cg_time = u'送货日期：2018 年 {month} 月 {day} 日'.format(month=month, day=cg_month_dict['day'])
    worksheet.cell(row=2, column=5, value=cg_time)

    month_dict_len = len(cg_month_dict['receiving_list'])

    print(cg_key, month, month_dict_len)

    # 删除多余空行
    worksheet.delete_rows(delivery_template_empty_row_index,
                          delivery_template_empty_row - month_dict_len - 2)  # 空余两行

    count_sum = '=SUM(C5:C{end_row})'.format(end_row=delivery_template_empty_row_index + month_dict_len)
    real_count_sum = '=SUM(D5:D{end_row})'.format(end_row=delivery_template_empty_row_index + month_dict_len)

    worksheet.cell(row=delivery_template_empty_row_index + month_dict_len + 2, column=3, value=count_sum)
    worksheet.cell(row=delivery_template_empty_row_index + month_dict_len + 2, column=4, value=real_count_sum)

    # 合并单元格
    worksheet.merge_cells(start_row=delivery_template_empty_row_index + month_dict_len + 6,
                          end_row=delivery_template_empty_row_index + month_dict_len + 6,
                          start_column=1,
                          end_column=3)

    worksheet.merge_cells(start_row=delivery_template_empty_row_index + month_dict_len + 6,
                          end_row=delivery_template_empty_row_index + month_dict_len + 6,
                          start_column=4,
                          end_column=6)
    worksheet.row_dimensions[delivery_template_empty_row_index + month_dict_len + 6].height = 38

    row = delivery_template_empty_row_index
    for sku_data in cg_month_dict['receiving_list']:
        worksheet.cell(row=row, column=1).value = sku_data["name"]
        worksheet.cell(row=row, column=2).value = sku_data["unit"]
        worksheet.cell(row=row, column=3).value = sku_data["count"]
        worksheet.cell(row=row, column=4).value = sku_data["real_count"]
        worksheet.cell(row=row, column=5).value = sku_data["sku"]
        worksheet.cell(row=row, column=6).value = sku_data["order"]
        row += 1

    # 删除多余空行
    worksheet.unmerge_cells(start_row=155, end_row=155, start_column=1, end_column=3)
    worksheet.unmerge_cells(start_row=155, end_row=155, start_column=4, end_column=5)
    worksheet.unmerge_cells(start_row=156, end_row=156, start_column=1, end_column=3)
    worksheet.unmerge_cells(start_row=157, end_row=157, start_column=1, end_column=3)
    worksheet.unmerge_cells(start_row=157, end_row=157, start_column=4, end_column=6)

    workbook.save(u'./output/delivery/{cg}送货单{month}月.xlsx'.format(cg=cg_key, month=month))



def test():
    pass

if __name__ == '__main__':

    export_invoice()
    export_delivery()
    # test()
    pass